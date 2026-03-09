
#!/usr/bin/env python3
"""
Local-ready Otodom scraper.

What it does
- walks Otodom apartment sale result pages
- extracts listing URLs from search pages
- opens each listing page
- saves flat-sale offers to CSV and XLSX

Examples
python otodom_scraper_local.py \
  --search-url "https://www.otodom.pl/pl/oferty/sprzedaz/mieszkanie/warszawa" \
  --max-pages 50 \
  --output-prefix otodom_warszawa

python otodom_scraper_local.py \
  --search-url "https://www.otodom.pl/pl/wyniki/sprzedaz/mieszkanie/mazowieckie/warszawa" \
  --start-page 1 \
  --max-pages 200 \
  --sleep 1.2

Notes
- This script is meant to be run on your own machine/network.
- Otodom changes markup often. If selectors move, update the constants below.
- Respect the site's robots.txt, terms, and rate limits.
"""
from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
import time
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional
from urllib.parse import urljoin, urlparse, parse_qs, urlencode, urlunparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


SEARCH_LINK_SELECTOR = 'a[data-cy="listing-item-link"]'
DESCRIPTION_SELECTOR = 'div[data-cy="adPageAdDescription"]'
TOP_TABLE_SELECTOR = 'div[data-testid="ad.top-information.table"]'
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/131.0.0.0 Safari/537.36"
)

FIELDNAMES = [
    "scraped_at_utc",
    "search_url",
    "listing_id",
    "title",
    "listing_url",
    "price_pln",
    "price_currency",
    "price_per_m2",
    "area_m2",
    "rooms",
    "market",
    "seller_type",
    "location_text",
    "street",
    "district",
    "city",
    "province",
    "latitude",
    "longitude",
    "floor",
    "building_year",
    "ownership",
    "heating",
    "rent_fee_pln",
    "balcony",
    "parking",
    "elevator",
    "description",
    "images_json",
    "attributes_json",
    "source_url",
]


@dataclass
class Offer:
    scraped_at_utc: str
    search_url: str
    listing_id: str = ""
    title: str = ""
    listing_url: str = ""
    price_pln: Optional[float] = None
    price_currency: str = "PLN"
    price_per_m2: Optional[float] = None
    area_m2: Optional[float] = None
    rooms: Optional[float] = None
    market: str = ""
    seller_type: str = ""
    location_text: str = ""
    street: str = ""
    district: str = ""
    city: str = ""
    province: str = ""
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    floor: str = ""
    building_year: str = ""
    ownership: str = ""
    heating: str = ""
    rent_fee_pln: Optional[float] = None
    balcony: str = ""
    parking: str = ""
    elevator: str = ""
    description: str = ""
    images_json: str = "[]"
    attributes_json: str = "{}"
    source_url: str = "https://www.otodom.pl/"


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": USER_AGENT,
        "Accept-Language": "pl-PL,pl;q=0.9,en;q=0.8",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    })
    return s


def clean_text(value: Optional[str]) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", " ", value).strip()


def parse_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(str(value))
    if not text:
        return None
    text = text.replace("\xa0", " ")
    # keep digits, minus, comma, dot
    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text:
        return None
    if text.count(",") and text.count("."):
        text = text.replace(" ", "")
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif text.count(",") and not text.count("."):
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def parse_intish(value: Any) -> str:
    n = parse_number(value)
    if n is None:
        return ""
    if abs(n - round(n)) < 1e-9:
        return str(int(round(n)))
    return str(n)


def absolute_url(base_url: str, href: str) -> str:
    return urljoin(base_url, href)


def update_query_param(url: str, name: str, value: Any) -> str:
    parsed = urlparse(url)
    query = parse_qs(parsed.query, keep_blank_values=True)
    query[name] = [str(value)]
    new_query = urlencode(query, doseq=True)
    return urlunparse(parsed._replace(query=new_query))


def get_html(session: requests.Session, url: str, timeout: int = 45) -> str:
    response = session.get(url, timeout=timeout)
    response.raise_for_status()
    return response.text


def soup_from_html(html: str) -> BeautifulSoup:
    return BeautifulSoup(html, "html.parser")


def find_json_blocks(soup: BeautifulSoup) -> List[Any]:
    payloads = []
    for script in soup.find_all("script"):
        text = script.string or script.get_text(" ", strip=True)
        text = text.strip() if text else ""
        if not text:
            continue
        if script.get("type") in {"application/ld+json", "application/json"} or text.startswith("{") or text.startswith("["):
            try:
                payloads.append(json.loads(text))
            except Exception:
                continue
    return payloads


def walk_json(node: Any) -> Iterable[Any]:
    yield node
    if isinstance(node, dict):
        for v in node.values():
            yield from walk_json(v)
    elif isinstance(node, list):
        for v in node:
            yield from walk_json(v)


def extract_listing_links(search_soup: BeautifulSoup, base_url: str) -> List[str]:
    links: List[str] = []
    seen = set()

    for a in search_soup.select(SEARCH_LINK_SELECTOR):
        href = a.get("href")
        if not href:
            continue
        full = absolute_url(base_url, href)
        if "/pl/oferta/" not in full and "/oferta/" not in full:
            continue
        if full not in seen:
            seen.add(full)
            links.append(full)

    # fallback: inspect embedded json for offer-like URLs
    if not links:
        for payload in find_json_blocks(search_soup):
            for node in walk_json(payload):
                if isinstance(node, dict):
                    url_val = node.get("url") or node.get("href")
                    if isinstance(url_val, str) and ("otodom.pl" in url_val or url_val.startswith("/")):
                        full = absolute_url(base_url, url_val)
                        if "/pl/oferta/" in full or "/oferta/" in full:
                            if full not in seen:
                                seen.add(full)
                                links.append(full)

    return links


def extract_ld_json_offer(soup: BeautifulSoup) -> Dict[str, Any]:
    for payload in find_json_blocks(soup):
        for node in walk_json(payload):
            if isinstance(node, dict):
                typ = node.get("@type")
                if typ in {"Product", "Offer", "House", "Apartment", "Residence"}:
                    return node
    return {}


def extract_attributes_table(soup: BeautifulSoup) -> Dict[str, str]:
    result: Dict[str, str] = {}
    container = soup.select_one(TOP_TABLE_SELECTOR)
    if container:
        texts = [clean_text(x.get_text(" ", strip=True)) for x in container.find_all(["div", "span", "p", "li"]) if clean_text(x.get_text(" ", strip=True))]
        # pair consecutive label/value items when possible
        for i in range(0, len(texts) - 1, 2):
            label = texts[i]
            value = texts[i + 1]
            if label and value and label not in result:
                result[label] = value

    # fallback: scan repeated label/value patterns across the page
    if not result:
        labels = soup.find_all(attrs={"data-cy": re.compile(r"table-label")})
        for label_el in labels:
            label = clean_text(label_el.get_text(" ", strip=True))
            value = ""
            sib = label_el.find_next(attrs={"data-testid": re.compile(r"table-value")})
            if sib:
                value = clean_text(sib.get_text(" ", strip=True))
            if label and value:
                result[label] = value
    return result


def lookup_attr(attrs: Dict[str, str], *patterns: str) -> str:
    norm = {clean_text(k).casefold(): v for k, v in attrs.items()}
    for pat in patterns:
        p = pat.casefold()
        for k, v in norm.items():
            if p in k:
                return v
    return ""


def extract_location_parts(location_text: str) -> Dict[str, str]:
    parts = [clean_text(x) for x in location_text.split(",") if clean_text(x)]
    out = {"street": "", "district": "", "city": "", "province": ""}
    if not parts:
        return out
    if len(parts) >= 1:
        out["street"] = parts[0]
    if len(parts) >= 2:
        out["district"] = parts[1]
    if len(parts) >= 3:
        out["city"] = parts[2]
    if len(parts) >= 4:
        out["province"] = parts[3]
    return out


def parse_offer_page(session: requests.Session, offer_url: str, search_url: str, sleep_s: float = 1.0) -> Offer:
    html = get_html(session, offer_url)
    soup = soup_from_html(html)
    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()

    title = ""
    h1 = soup.find("h1")
    if h1:
        title = clean_text(h1.get_text(" ", strip=True))

    description = ""
    desc = soup.select_one(DESCRIPTION_SELECTOR)
    if desc:
        description = clean_text(desc.get_text(" ", strip=True))

    attrs = extract_attributes_table(soup)
    location_text = ""

    # Attempt to find address/location text from visible content or JSON-LD
    ld_offer = extract_ld_json_offer(soup)
    geo = {}
    images = []

    if isinstance(ld_offer, dict):
        title = title or clean_text(ld_offer.get("name", ""))
        offer_id = clean_text(str(ld_offer.get("sku") or ld_offer.get("productID") or ld_offer.get("@id") or ""))
        offers_obj = ld_offer.get("offers") if isinstance(ld_offer.get("offers"), dict) else {}
        price_val = parse_number(offers_obj.get("price"))
        currency = clean_text(offers_obj.get("priceCurrency") or "PLN") or "PLN"
        geo = ld_offer.get("geo") if isinstance(ld_offer.get("geo"), dict) else {}
        address = ld_offer.get("address")
        if isinstance(address, dict):
            location_bits = [
                clean_text(address.get("streetAddress")),
                clean_text(address.get("addressLocality")),
                clean_text(address.get("addressRegion")),
            ]
            location_text = ", ".join([x for x in location_bits if x])
        imgs = ld_offer.get("image") or ld_offer.get("images") or []
        if isinstance(imgs, str):
            images = [imgs]
        elif isinstance(imgs, list):
            images = [str(x) for x in imgs if x]
    else:
        offer_id = ""

    # visible location fallback
    if not location_text:
        loc_candidates = []
        for tag in soup.find_all(["a", "span", "p", "div"]):
            text = clean_text(tag.get_text(" ", strip=True))
            if not text:
                continue
            if 5 <= len(text) <= 140 and text.count(",") >= 1:
                loc_candidates.append(text)
        if loc_candidates:
            location_text = loc_candidates[0]

    location_parts = extract_location_parts(location_text)
    area = parse_number(
        lookup_attr(attrs, "powierzchnia", "area")
    )
    rooms = parse_number(
        lookup_attr(attrs, "liczba pokoi", "rooms")
    )
    floor = lookup_attr(attrs, "piętro", "floor")
    building_year = lookup_attr(attrs, "rok budowy", "year")
    ownership = lookup_attr(attrs, "forma własności", "ownership")
    heating = lookup_attr(attrs, "ogrzewanie", "heating")
    rent_fee = parse_number(lookup_attr(attrs, "czynsz", "rent"))

    market = lookup_attr(attrs, "rynek", "market")
    seller_type = lookup_attr(attrs, "ogłoszeniodawca", "sprzedawca", "seller")

    if not title:
        title_tag = soup.find("title")
        title = clean_text(title_tag.get_text(" ", strip=True)) if title_tag else ""

    # price fallbacks
    if 'price_val' not in locals():
        price_val = None
    if 'currency' not in locals():
        currency = "PLN"

    if price_val is None:
        full_text = clean_text(soup.get_text(" ", strip=True))
        m = re.search(r'([0-9][0-9\s\xa0.,]{3,})\s*zł', full_text, re.IGNORECASE)
        if m:
            price_val = parse_number(m.group(1))

    price_per_m2 = None
    if price_val is not None and area:
        try:
            price_per_m2 = round(price_val / area, 2)
        except Exception:
            price_per_m2 = None

    lower_blob = (json.dumps(attrs, ensure_ascii=False) + " " + description).casefold()
    balcony = "yes" if "balkon" in lower_blob else ""
    parking = "yes" if ("garaż" in lower_blob or "miejsce parkingowe" in lower_blob or "parking" in lower_blob) else ""
    elevator = "yes" if ("winda" in lower_blob or "elevator" in lower_blob) else ""

    offer = Offer(
        scraped_at_utc=now,
        search_url=search_url,
        listing_id=offer_id,
        title=title,
        listing_url=offer_url,
        price_pln=price_val,
        price_currency=currency,
        price_per_m2=price_per_m2,
        area_m2=area,
        rooms=rooms,
        market=market,
        seller_type=seller_type,
        location_text=location_text,
        street=location_parts["street"],
        district=location_parts["district"],
        city=location_parts["city"],
        province=location_parts["province"],
        latitude=parse_number(geo.get("latitude")) if isinstance(geo, dict) else None,
        longitude=parse_number(geo.get("longitude")) if isinstance(geo, dict) else None,
        floor=floor,
        building_year=building_year,
        ownership=ownership,
        heating=heating,
        rent_fee_pln=rent_fee,
        balcony=balcony,
        parking=parking,
        elevator=elevator,
        description=description,
        images_json=json.dumps(images, ensure_ascii=False),
        attributes_json=json.dumps(attrs, ensure_ascii=False),
        source_url="https://www.otodom.pl/",
    )
    time.sleep(max(0.0, sleep_s))
    return offer


def scrape_search(session: requests.Session, search_url: str, start_page: int, max_pages: int, sleep_s: float) -> List[Offer]:
    offers: List[Offer] = []
    seen_listing_urls = set()

    for page in range(start_page, start_page + max_pages):
        page_url = update_query_param(search_url, "page", page)
        print(f"[search] page={page} url={page_url}", file=sys.stderr)
        try:
            html = get_html(session, page_url)
        except Exception as e:
            print(f"[warn] search page failed: {page_url} :: {e}", file=sys.stderr)
            break

        soup = soup_from_html(html)
        listing_urls = extract_listing_links(soup, page_url)
        if not listing_urls:
            print(f"[stop] no listing links found on page {page}", file=sys.stderr)
            break

        new_count = 0
        for listing_url in listing_urls:
            if listing_url in seen_listing_urls:
                continue
            seen_listing_urls.add(listing_url)
            new_count += 1
            print(f"[offer] {listing_url}", file=sys.stderr)
            try:
                offers.append(parse_offer_page(session, listing_url, search_url, sleep_s=sleep_s))
            except Exception as e:
                print(f"[warn] offer failed: {listing_url} :: {e}", file=sys.stderr)

        if new_count == 0:
            print(f"[stop] page {page} yielded only duplicates", file=sys.stderr)
            break

        time.sleep(max(0.0, sleep_s))

    return offers


def write_csv(path: str, offers: List[Offer]) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        for offer in offers:
            row = asdict(offer)
            writer.writerow(row)


def auto_width(ws) -> None:
    for col_cells in ws.columns:
        values = [clean_text(str(c.value)) for c in col_cells if c.value is not None]
        max_len = max((len(v) for v in values), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 10), 40)


def write_xlsx(path: str, offers: List[Offer]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "offers"
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E1F2")

    ws.append(FIELDNAMES)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=thin)

    for offer in offers:
        ws.append([asdict(offer).get(k) for k in FIELDNAMES])

    if offers:
        end_row = len(offers) + 1
        last_col = get_column_letter(len(FIELDNAMES))
        table = Table(displayName="OtodomOffers", ref=f"A1:{last_col}{end_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    # Number formats
    num_cols = {
        "F": '#,##0.00',
        "H": '#,##0.00',
        "I": '#,##0.00',
        "J": '0.0',
        "R": '#,##0.00',
    }
    for col, fmt in num_cols.items():
        for cell in ws[col][1:]:
            cell.number_format = fmt
            cell.alignment = Alignment(vertical="center")

    for col in ["B", "D", "E", "L", "AB", "AC"]:
        for cell in ws[col][1:]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    auto_width(ws)

    notes = wb.create_sheet("notes")
    notes["A1"] = "How to use"
    notes["A1"].font = Font(bold=True, size=14)
    notes["A3"] = "1) Run the scraper locally."
    notes["A4"] = '2) Open the produced XLSX or import the CSV into Google Sheets.'
    notes["A5"] = "3) If selectors break, update SEARCH_LINK_SELECTOR / DESCRIPTION_SELECTOR / TOP_TABLE_SELECTOR in the script."
    notes["A7"] = "Source URL pattern"
    notes["A8"] = "Use an Otodom apartment-sale results URL, for example:"
    notes["A9"] = "https://www.otodom.pl/pl/oferty/sprzedaz/mieszkanie/warszawa"
    notes["A10"] = "or"
    notes["A11"] = "https://www.otodom.pl/pl/wyniki/sprzedaz/mieszkanie/mazowieckie/warszawa"
    notes.column_dimensions["A"].width = 110

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Scrape Otodom apartment sale offers to CSV/XLSX.")
    parser.add_argument("--search-url", required=True, help="Otodom apartment-sale search results URL.")
    parser.add_argument("--start-page", type=int, default=1, help="Page number to start from.")
    parser.add_argument("--max-pages", type=int, default=10, help="How many pages to walk.")
    parser.add_argument("--sleep", type=float, default=1.0, help="Delay between requests in seconds.")
    parser.add_argument("--output-prefix", default="otodom_offers", help="Output file prefix.")
    args = parser.parse_args()

    session = make_session()
    offers = scrape_search(
        session=session,
        search_url=args.search_url,
        start_page=args.start_page,
        max_pages=args.max_pages,
        sleep_s=args.sleep,
    )

    if not offers:
        print("No offers were collected. Check the URL, selectors, or network access.", file=sys.stderr)
        return 2

    csv_path = f"{args.output_prefix}.csv"
    xlsx_path = f"{args.output_prefix}.xlsx"

    write_csv(csv_path, offers)
    write_xlsx(xlsx_path, offers)

    print(json.dumps({
        "offers": len(offers),
        "csv": csv_path,
        "xlsx": xlsx_path,
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
